"""
AI Bongchae 뉴스 엑셀 → Document(source="뉴스") 임포트.

엑셀 컬럼(News_List 시트):
  A 제목 | B 카테고리 | C 핵심 요약 | D 출처 URL | E 작성 엔진 | F 등록 시간

요약(C)은 이미 Gemini/Claude로 완료돼 있으므로 EXAONE 재요약 없이 그대로 저장한다.
등록 시간("2026. 6. 12. 오전 11:07:22")을 파싱해 카드 표시 시각(created_at)으로 넣는다.
중복은 url 로 판정한다.

사용 예:
    python manage.py import_news_excel
    python manage.py import_news_excel --file "C:/path/AI_Bongchae_News.xlsx"
    python manage.py import_news_excel --dry-run        # 저장 없이 미리보기
"""
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.utils import timezone

from insight.models import Source, Document

DEFAULT_FILE = (Path(settings.BASE_DIR).parent
                / "data_source" / "scripts" / "AI_Bongchae_News_2026-06-14.xlsx")

# "2026. 6. 12. 오전 11:07:22"
_DT = re.compile(r"(\d+)\.\s*(\d+)\.\s*(\d+)\.\s*(오전|오후)\s*(\d+):(\d+):(\d+)")


def parse_korean_dt(s):
    """한국어 등록시간 문자열 → aware datetime (실패 시 None)."""
    if not s:
        return None
    m = _DT.search(str(s))
    if not m:
        return None
    y, mo, d, ampm, hh, mi, se = (int(m.group(i)) if i != 4 else m.group(i)
                                  for i in range(1, 8))
    if ampm == "오후" and hh != 12:
        hh += 12
    elif ampm == "오전" and hh == 12:
        hh = 0
    try:
        naive = datetime(y, mo, d, hh, mi, se)
    except ValueError:
        return None
    if settings.USE_TZ:
        return timezone.make_aware(naive, timezone.get_default_timezone())
    return naive


def media_from_url(url):
    """URL 도메인에서 매체명 추출 (www 제거). 예: aitimes.com"""
    try:
        host = urlparse(url).netloc
        return host[4:] if host.startswith("www.") else host
    except Exception:  # noqa: BLE001
        return ""


class Command(BaseCommand):
    help = "AI Bongchae 뉴스 엑셀을 읽어 Document(뉴스)로 저장한다."

    def add_arguments(self, parser):
        parser.add_argument("--file", type=str, default=str(DEFAULT_FILE),
                            help="엑셀 파일 경로")
        parser.add_argument("--dry-run", action="store_true",
                            help="저장하지 않고 파싱 결과만 미리보기")

    def handle(self, *args, **opts):
        path = Path(opts["file"])
        if not path.exists():
            self.stderr.write(self.style.ERROR(f"파일 없음: {path}"))
            return
        dry = opts["dry_run"]

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["News_List"] if "News_List" in wb.sheetnames else wb[wb.sheetnames[0]]

        source, _ = Source.objects.get_or_create(
            name="뉴스", defaults={"type": Source.Type.NEWS})

        n_new = n_dup = n_skip = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            title, category, summary, url, engine, reg = (list(row) + [None] * 6)[:6]
            title = (title or "").strip()
            url = (url or "").strip()
            summary = (summary or "").strip()
            if not title or not url:
                n_skip += 1
                continue
            if Document.objects.filter(url=url).exists():
                n_dup += 1
                continue

            dt = parse_korean_dt(reg)
            pub = dt.date() if dt else None
            if dry:
                if n_new < 3:
                    self.stdout.write(
                        f"[{category}] {title[:40]} | {engine} | {dt} | {url[:40]}")
                n_new += 1
                continue

            doc = Document.objects.create(
                source=source,
                title=title[:500],
                authors=media_from_url(url),
                published_date=pub,
                raw_text=summary,
                summary=summary,
                category=(category or "기타").strip(),
                engine=(engine or "").strip(),
                url=url,
                status=Document.Status.ANALYZED,
            )
            # created_at 은 auto_now_add 라 create 시 무시됨 → update 로 등록시간 반영
            if dt:
                Document.objects.filter(pk=doc.pk).update(created_at=dt)
            n_new += 1

        verb = "미리보기" if dry else "저장"
        self.stdout.write(self.style.SUCCESS(
            f"\n[뉴스 엑셀 {verb} 완료] 신규 {n_new} / 중복 {n_dup} / 건너뜀 {n_skip}"))
